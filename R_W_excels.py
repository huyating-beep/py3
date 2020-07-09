# encoding: utf-8 
# @time: 2020/7/2 2:33 下午
# @author: hyt
# @contact: 2621408918@qq.com
# 导入读写
from openpyxl import load_workbook
from openpyxl import Workbook


# 读数据
def read_data(wb_name, sheet_name):
    wb_data = []
    wb_name = load_workbook(wb_name)
    sheet_name = wb_name[sheet_name]
    # 获取最大行
    row_all = sheet_name.max_row
    column_all = sheet_name.max_column
    # 获取最大列
    for i in range(2, row_all + 1):
        rows = []
        for j in range(1, column_all - 1):
            wb_all = sheet_name.cell(row=i, column=j).value
            rows.append(wb_all)
        wb_data.append(rows)
    return wb_data
    wb_name.save("test_py.xlsx")


# 写数据
def wite_data(wb_name, sheet_namee, row, column, value):
    wis = load_workbook(wb_name)
    wis_sheetname = wis[sheet_namee]
    wis_sheetname.cell(row=row, column=column).value = value
    wis.save("test_py.xlsx")

if __name__ == '__main__':
    # 调用函数
    read_data('test_py.xlsx', 'info_test')
